# components/tab_export.py

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta 
import json
import shutil
import pandas as pd
import json
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage
from dotenv import load_dotenv
import os
from functools import lru_cache
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import pytz

load_dotenv()

# ìºì‹œëœ í”„ë¡œì íŠ¸ ë¡œë”©
@st.cache_data(ttl=300)  # 5ë¶„ TTL
def _load_all_histories():
    """ëª¨ë“  ëŒ€í™” ê¸°ë¡ì„ ìºì‹œì™€ í•¨ê»˜ ë¡œë“œ"""
    try:
        if not os.path.exists("chat_histories.json"):
            return {}
        with open("chat_histories.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        st.error(f"íŒŒì¼ ë¡œë“œ ì‹¤íŒ¨: {e}")
        return {}

def show_export_helper():
    """ìˆ˜ì¶œ ì œì•ˆì„œ ë„ìš°ë¯¸ ë©”ì¸ í•¨ìˆ˜ - ìµœì í™” ë²„ì „"""

    # ì•ˆë‚´ ë©”ì‹œì§€
    st.info("""
    **ì‚¬ìš© ë°©ë²•:**
    ë¯¸êµ­ ì‹œì¥ ìˆ˜ì¶œìš© ìƒí’ˆ ê¸°íš ë‹¨ê³„ì—ì„œ í™œìš©í•  ìˆ˜ ìˆëŠ” ë¬¸ì„œ ì‘ì„± ë„ìš°ë¯¸ì…ë‹ˆë‹¤.

    1. ì±—ë´‡ê³¼ ì§ˆì˜ì‘ë‹µ ì‹œ ì„¤ì •í•œ í”„ë¡œì íŠ¸ëª…ì„ ì„ íƒí•˜ì„¸ìš”.
    2. ì¢Œì¸¡ ë¹ˆ ì¹¸ì— ì œí’ˆ ì •ë³´ ë° ì œì•ˆ ì˜ë„ë¥¼ ì…ë ¥í•˜ë©´ ë¬¸ì„œë¡œ ìë™ ì €ì¥ë©ë‹ˆë‹¤.
    3. ìš°ì¸¡ í•˜ë‹¨ ë²„íŠ¼ì„ í´ë¦­í•˜ë©´ ì •ë¦¬ëœ ë¬¸ì„œë¡œ ì¶œë ¥ì´ ê°€ëŠ¥í•©ë‹ˆë‹¤.
    """)
    
    # ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™” - ì¡°ê±´ë¶€
    init_session_state()

    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("**1ï¸âƒ£ ì •ë³´ ì…ë ¥**")
        show_basic_info_form()
    
    with col2:
        st.markdown("**2ï¸âƒ£ ë¶„ì„ ë¦¬í¬íŠ¸ ì˜ˆì‹œì‚¬ì§„**")
        render_guide_section()

def init_session_state():
    """ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™” - ìµœì í™”"""
    defaults = {
        "export_data": {},
        "report_generated": False,
        "selected_template_data": "ê¸°ë³¸ ì œì•ˆì„œ",
        "show_summary_area": False,
        "summary_content": "",
        "ai_processing": False  # AI ì²˜ë¦¬ ìƒíƒœ ì¶”ê°€
    }
    
    for key, default_value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default_value

def render_guide_section():
    """ê°€ì´ë“œ ì´ë¯¸ì§€ ë° ì—‘ì…€ ë²„íŠ¼ ì„¹ì…˜"""
    try:
        st.image('./ê°€ì´ë“œ.png')
    except FileNotFoundError:
        st.warning("ì´ë¯¸ì§€ íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
    except Exception as e:
        st.error(f"ì´ë¯¸ì§€ ë¡œë“œ ì˜¤ë¥˜: {e}")
    
    st.markdown("---")
    add_excel_export_button()

def get_available_projects():
    """ì €ì¥ëœ í”„ë¡œì íŠ¸ ëª©ë¡ì„ ê°€ì ¸ì˜¤ëŠ” í•¨ìˆ˜ - í”„ë¡œì íŠ¸ëª…ë§Œ ì¶”ì¶œ"""
    try:
        all_histories = _load_all_histories()
        
        # í”„ë¡œì íŠ¸ëª…ë§Œ ì¶”ì¶œ (ëª¨ë“œ ë¶€ë¶„ ì œê±°)
        project_names = set()
        for project_key in all_histories.keys():
            # ë§ˆì§€ë§‰ ì–¸ë”ìŠ¤ì½”ì–´ë¡œ êµ¬ë¶„í•˜ì—¬ í”„ë¡œì íŠ¸ëª… ì¶”ì¶œ
            if '_' in project_key:
                # ë§ˆì§€ë§‰ '_'ë¥¼ ê¸°ì¤€ìœ¼ë¡œ ë¶„ë¦¬
                parts = project_key.rsplit('_', 1)
                if len(parts) == 2 and parts[1] in ['ê·œì œ', 'ë¦¬ì½œì‚¬ë¡€']:
                    project_names.add(parts[0])
            else:
                # ì–¸ë”ìŠ¤ì½”ì–´ê°€ ì—†ëŠ” ê²½ìš° ì „ì²´ë¥¼ í”„ë¡œì íŠ¸ëª…ìœ¼ë¡œ ê°„ì£¼
                project_names.add(project_key)
        
        return sorted(list(project_names))
    except Exception as e:
        st.error(f"í”„ë¡œì íŠ¸ ëª©ë¡ ë¶ˆëŸ¬ì˜¤ê¸° ì‹¤íŒ¨: {e}")
        return []

def load_project_chat_history(project_name):
    """íŠ¹ì • í”„ë¡œì íŠ¸ì˜ í†µí•© ì±„íŒ… íˆìŠ¤í† ë¦¬ ë¶ˆëŸ¬ì˜¤ê¸°"""
    try:
        all_histories = _load_all_histories()
        
        # í•´ë‹¹ í”„ë¡œì íŠ¸ì˜ ëª¨ë“  ëª¨ë“œ ë°ì´í„° ìˆ˜ì§‘
        regulation_history = []
        recall_history = []
        
        # ê·œì œ ëª¨ë“œ ë°ì´í„°
        regulation_key = f"{project_name}_ê·œì œ"
        if regulation_key in all_histories:
            regulation_data = all_histories[regulation_key]
            regulation_history = regulation_data.get("chat_history", [])
        
        # ë¦¬ì½œì‚¬ë¡€ ëª¨ë“œ ë°ì´í„°
        recall_key = f"{project_name}_ë¦¬ì½œì‚¬ë¡€"
        if recall_key in all_histories:
            recall_data = all_histories[recall_key]
            recall_history = recall_data.get("chat_history", [])
        
        # ë‘ íˆìŠ¤í† ë¦¬ë¥¼ í•©ì³ì„œ ë°˜í™˜ (ì‹œê°„ìˆœ ì •ë ¬ ê°€ëŠ¥)
        combined_history = regulation_history + recall_history
        
        return combined_history
        
    except Exception as e:
        st.error(f"í”„ë¡œì íŠ¸ íˆìŠ¤í† ë¦¬ ë¶ˆëŸ¬ì˜¤ê¸° ì‹¤íŒ¨: {e}")
        return []

def get_project_summary_info(project_name):
    """í”„ë¡œì íŠ¸ì˜ ìš”ì•½ ì •ë³´ ë°˜í™˜"""
    try:
        all_histories = _load_all_histories()
        
        regulation_key = f"{project_name}_ê·œì œ"
        recall_key = f"{project_name}_ë¦¬ì½œì‚¬ë¡€"
        
        info = {
            "regulation_chats": 0,
            "recall_chats": 0,
            "last_updated": None,
            "modes": []
        }
        
        if regulation_key in all_histories:
            reg_data = all_histories[regulation_key]
            info["regulation_chats"] = len(reg_data.get("chat_history", [])) // 2
            info["modes"].append("ê·œì œ")
            if reg_data.get("last_updated"):
                info["last_updated"] = reg_data["last_updated"]
        
        if recall_key in all_histories:
            recall_data = all_histories[recall_key]
            info["recall_chats"] = len(recall_data.get("chat_history", [])) // 2
            info["modes"].append("ë¦¬ì½œì‚¬ë¡€")
            # ë” ìµœê·¼ ì—…ë°ì´íŠ¸ ì‹œê°„ ì„ íƒ
            if recall_data.get("last_updated"):
                if not info["last_updated"] or recall_data["last_updated"] > info["last_updated"]:
                    info["last_updated"] = recall_data["last_updated"]
        
        return info
        
    except Exception as e:
        return {"regulation_chats": 0, "recall_chats": 0, "last_updated": None, "modes": []}

def show_basic_info_form():
    """ê¸°ë³¸ ì •ë³´ ì…ë ¥ í¼ - ìµœì í™”"""
    narrow_col, _ = st.columns([0.8, 0.2])

    with narrow_col:
        render_project_selector()
        st.markdown("---")
        render_product_info_section()
        render_background_section()
        render_risk_summary_section()

def render_project_selector():
    """í”„ë¡œì íŠ¸ ì„ íƒ ì„¹ì…˜"""
    st.markdown("**í”„ë¡œì íŠ¸ ì„ íƒ**")
    
    available_projects = get_available_projects()
    
    if available_projects:
        selected_project = st.selectbox(
            "ì €ì¥ëœ í”„ë¡œì íŠ¸ì—ì„œ ì„ íƒ",
            ["ìƒˆ í”„ë¡œì íŠ¸"] + available_projects,
            key="project_selector",
            help="ê¸°ì¡´ í”„ë¡œì íŠ¸ë¥¼ ì„ íƒí•˜ì—¬ ê·œì œ/ë¦¬ì½œì‚¬ë¡€ ëª¨ë“  Q&A ê¸°ë¡ì„ í†µí•©í•˜ì—¬ ë¶ˆëŸ¬ì˜µë‹ˆë‹¤."
        )
        
        if selected_project != "ìƒˆ í”„ë¡œì íŠ¸":
            # í”„ë¡œì íŠ¸ ì •ë³´ í‘œì‹œ - ì»¬ëŸ¼ ì¤‘ì²© ë°©ì§€
            project_info = get_project_summary_info(selected_project)
            
            # ì„¸ë¡œë¡œ ë°°ì¹˜í•˜ì—¬ ì»¬ëŸ¼ ì¤‘ì²© ë°©ì§€
            # st.write(f"**ê·œì œ ëŒ€í™”:** {project_info['regulation_chats']}ê±´")
            # st.write(f"**ë¦¬ì½œ ëŒ€í™”:** {project_info['recall_chats']}ê±´")
            # total_chats = project_info['regulation_chats'] + project_info['recall_chats']
            # st.write(f"**ì´ ëŒ€í™”:** {total_chats}ê±´")
            
            # if project_info['modes']:
            #     st.info(f"**ì„ íƒëœ í”„ë¡œì íŠ¸:** {selected_project}  \n**í¬í•¨ ëª¨ë“œ:** {', '.join(project_info['modes'])}")
    else:
        st.info("ì €ì¥ëœ í”„ë¡œì íŠ¸ê°€ ì—†ìŠµë‹ˆë‹¤. ì±„íŒ… íƒ­ì—ì„œ ëŒ€í™” í›„ ì €ì¥í•´ì£¼ì„¸ìš”.")
        selected_project = "ìƒˆ í”„ë¡œì íŠ¸"

def render_product_info_section():
    """ì œí’ˆ ì •ë³´ ì…ë ¥ ì„¹ì…˜"""
    st.markdown("**ì œí’ˆ ì •ë³´**")
    
    # ì»¬ëŸ¼ ì¤‘ì²© ë°©ì§€ - ì„¸ë¡œë¡œ ë°°ì¹˜
    product_name = st.text_input(
        "ì œí’ˆëª…", 
        placeholder="ë‹¨ë°±ì§ˆ ì—ë„ˆì§€ë°”", 
        key="product_name"
    )
    
    target_market = st.text_input(
        "íƒ€ê²Ÿì¸µ", 
        placeholder="30ëŒ€ ì—¬ì„±", 
        key="target_name"
    )

def render_background_section():
    """ì¶”ì§„ë°°ê²½ ì…ë ¥ ì„¹ì…˜"""
    st.markdown("**ì¶”ì§„ë°°ê²½**")
    
    # í”Œë ˆì´ìŠ¤í™€ë” í…ìŠ¤íŠ¸ ë‹¨ì¶•
    placeholder_text = """ì‹œì¥ ë¶„ì„, ê²½ìŸì‚¬ ë‚´ìš©ì„ ì…ë ¥í•˜ì„¸ìš”.

ì˜ˆì‹œ) ë¯¸êµ­ ë‚´ 30ëŒ€ ì—¬ì„±ì„ ì¤‘ì‹¬ìœ¼ë¡œ ê³ ë‹¨ë°± ì‹í’ˆì— ëŒ€í•œ ìˆ˜ìš”ê°€ í¬ê²Œ ëŠ˜ê³  ìˆìœ¼ë©°, 2022ë…„ë¶€í„° 2024ë…„ê¹Œì§€ ë‹¨ë°±ì§ˆ ê°„ì‹ì€ ì—°í‰ê·  9%ì˜ ì„±ì¥ë¥ ì„ ê¸°ë¡í•˜ê³  ìˆìŠµë‹ˆë‹¤...

(ìƒì„¸í•œ ì‹œì¥ ë¶„ì„ ë° ê²½ìŸì‚¬ ì •ë³´ ì…ë ¥)"""
    
    background = st.text_area(
        "ì œì•ˆ ì˜ë„",
        placeholder=placeholder_text,
        height=350,  # ë†’ì´ ì¡°ì •
        key="background"
    )

def render_risk_summary_section():
    """ê·œì œ ë¦¬ìŠ¤í¬ ìš”ì•½ ì„¹ì…˜"""
    st.markdown("**ê·œì œ ë¦¬ìŠ¤í¬ ìš”ì•½**")
    
    selected_project = st.session_state.get("project_selector", "ìƒˆ í”„ë¡œì íŠ¸")
    
    # ë²„íŠ¼ í…ìŠ¤íŠ¸ ë™ì  ìƒì„±
    if selected_project != "ìƒˆ í”„ë¡œì íŠ¸":
        project_info = get_project_summary_info(selected_project)
        total_chats = project_info['regulation_chats'] + project_info['recall_chats']
        button_text = f"'{selected_project}' í”„ë¡œì íŠ¸ Q&A ë¶„ì„"
    else:
        button_text = "í˜„ì¬ ì„¸ì…˜ Q&A ë‚´ìš© ë¶ˆëŸ¬ì˜¤ê¸°"

    # AI ì²˜ë¦¬ ìƒíƒœì— ë”°ë¥¸ ë²„íŠ¼ ë¹„í™œì„±í™”
    button_disabled = st.session_state.get("ai_processing", False)
    
    if st.button(button_text, disabled=button_disabled):
        process_qa_analysis(selected_project)

def process_qa_analysis(selected_project):
    """Q&A ë¶„ì„ ì²˜ë¦¬ - ë¶„ë¦¬ëœ í•¨ìˆ˜"""
    st.session_state.ai_processing = True
    st.session_state.show_summary_area = True
    
    try:
        # í”„ë¡œì íŠ¸ ë°ì´í„° ë¡œë“œ
        if selected_project != "ìƒˆ í”„ë¡œì íŠ¸":
            chat_history = load_project_chat_history(selected_project)
            if not chat_history:
                st.warning(f"'{selected_project}' í”„ë¡œì íŠ¸ì— ëŒ€í™” ê¸°ë¡ì´ ì—†ìŠµë‹ˆë‹¤.")
                return
        else:
            chat_history = st.session_state.get("chat_history", [])
        
        if not chat_history:
            st.warning("âš ï¸ ë¶ˆëŸ¬ì˜¬ ëŒ€í™” ê¸°ë¡ì´ ì—†ìŠµë‹ˆë‹¤. ë¨¼ì € ì±„íŒ… íƒ­ì—ì„œ ëŒ€í™”ë¥¼ ì§„í–‰í•´ì£¼ì„¸ìš”.")
            return
        
        # Q&A í…ìŠ¤íŠ¸ ìƒì„±
        qa_text = generate_qa_text(chat_history)
        
        if qa_text:
            # AI ë¶„ì„ ìˆ˜í–‰
            perform_ai_analysis(qa_text, selected_project)
        
    except Exception as e:
        st.error(f"âŒ ë¶„ì„ ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜: {e}")
    finally:
        st.session_state.ai_processing = False
        st.rerun()

def generate_qa_text(chat_history):
    """ì±„íŒ… íˆìŠ¤í† ë¦¬ì—ì„œ Q&A í…ìŠ¤íŠ¸ ìƒì„±"""
    qa_text = ""
    for i in range(0, len(chat_history), 2):
        if i + 1 < len(chat_history):
            question = chat_history[i]["content"]
            answer = chat_history[i + 1]["content"]
            qa_text += f"ì§ˆë¬¸: {question}\në‹µë³€: {answer}\n\n"
    return qa_text

@st.cache_data(ttl=1800)  # 30ë¶„ ìºì‹œ
def perform_ai_analysis_cached(qa_text, openai_api_key):
    """AI ë¶„ì„ ìˆ˜í–‰ - ìºì‹œ ì ìš©"""
    try:
        llm = ChatOpenAI(
            model="gpt-4o-mini", 
            temperature=0.3,
            openai_api_key=openai_api_key
        )
        
        # í†µí•© ë¶„ì„ í”„ë¡¬í”„íŠ¸ (ë‹¨ì¼ ìš”ì²­ìœ¼ë¡œ ìµœì í™”)
        analysis_prompt = f"""
ë‹¤ìŒ Q&A ëŒ€í™”ë“¤ì„ ë¶„ì„í•˜ì—¬ ê·œì œ ë° ë¦¬ì½œì‚¬ë¡€ ê´€ë ¨ ë‚´ìš©ì„ ìš”ì•½í•´ì£¼ì„¸ìš”.

ë¶„ì„ ìš”êµ¬ì‚¬í•­:
1. ê·œì œ ê´€ë ¨ ë‚´ìš© (FDA ê·œì •, ë²•ë ¹, í—ˆê°€, ë“±ë¡, ë¼ë²¨ë§ ë“±)
2. ë¦¬ì½œì‚¬ë¡€ ê´€ë ¨ ë‚´ìš© (ì œí’ˆ ë¦¬ì½œ, íšŒìˆ˜, ì•ˆì „ ê²½ê³  ë“±)

ê° ì¹´í…Œê³ ë¦¬ë³„ë¡œ 3-4ë¬¸ì¥ìœ¼ë¡œ í•µì‹¬ ë‚´ìš©ì„ ìš”ì•½í•˜ê³ , í•´ë‹¹ ë‚´ìš©ì´ ì—†ëŠ” ê²½ìš° "ê´€ë ¨ ë‚´ìš© ì—†ìŒ"ìœ¼ë¡œ í‘œì‹œí•´ì£¼ì„¸ìš”.

ì‘ë‹µ í˜•ì‹:
ğŸ“‹ **ê·œì œ ê´€ë ¨ ìš”ì•½**
[ê·œì œ ê´€ë ¨ ìš”ì•½ ë‚´ìš©]

ğŸš¨ **ë¦¬ì½œì‚¬ë¡€ ìš”ì•½**
[ë¦¬ì½œì‚¬ë¡€ ê´€ë ¨ ìš”ì•½ ë‚´ìš©]

Q&A ë‚´ìš©:
{qa_text}
"""
        
        response = llm.invoke([HumanMessage(content=analysis_prompt)])
        final_summary = response.content.strip()
        
        # URL ë° ë¶ˆí•„ìš”í•œ ë‚´ìš© ì œê±°
        import re
        final_summary = re.sub(r'https?://[^\s]+', '', final_summary)
        final_summary = re.sub(r'ğŸ“.*?ì¶œì²˜:.*', '', final_summary, flags=re.DOTALL)
        
        return final_summary
        
    except Exception as e:
        return f"AI ë¶„ì„ ì‹¤íŒ¨: {str(e)}"

def perform_ai_analysis(qa_text, selected_project):
    """AI ë¶„ì„ ìˆ˜í–‰"""
    with st.spinner("ğŸ¤– AIê°€ ëŒ€í™” ë‚´ìš©ì„ í†µí•© ë¶„ì„í•˜ê³  ìˆìŠµë‹ˆë‹¤..."):
        try:
            openai_api_key = os.getenv("OPENAI_API_KEY")
            if not openai_api_key:
                st.error("OpenAI API í‚¤ê°€ ì„¤ì •ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.")
                return
            
            # ìºì‹œëœ AI ë¶„ì„ ìˆ˜í–‰
            final_summary = perform_ai_analysis_cached(qa_text, openai_api_key)
            st.session_state.summary_content = final_summary
            
            # ì„±ê³µ ë©”ì‹œì§€
            if selected_project != "ìƒˆ í”„ë¡œì íŠ¸":
                project_info = get_project_summary_info(selected_project)
                total_chats = project_info['regulation_chats'] + project_info['recall_chats']
                st.success(f"âœ… '{selected_project}' í”„ë¡œì íŠ¸ì˜ {total_chats}ê±´ Q&Aë¥¼ ì„±ê³µì ìœ¼ë¡œ ë¶„ì„í–ˆìŠµë‹ˆë‹¤!")
            else:
                st.success("âœ… í˜„ì¬ ì„¸ì…˜ì˜ Q&Aë¥¼ ì„±ê³µì ìœ¼ë¡œ ë¶„ì„í–ˆìŠµë‹ˆë‹¤!")
                
        except Exception as e:
            st.error(f"âŒ AI ë¶„ì„ ì¤‘ ì˜¤ë¥˜: {e}")
            st.session_state.summary_content = f"ë¶„ì„ ì‹¤íŒ¨: {e}"

# ìš”ì•½ ë‚´ìš© í‘œì‹œ ì„¹ì…˜ì„ show_basic_info_form() ëì— ì¶”ê°€
def render_summary_display():
    """ìš”ì•½ ë‚´ìš© í‘œì‹œ"""
    if st.session_state.get("show_summary_area", False):
        st.markdown("### ğŸ“Š í†µí•© ëŒ€í™” ë¶„ì„ ê²°ê³¼")
        
        edited_summary = st.text_area(
            "ğŸ“ ê·œì œ/ë¦¬ì½œ í†µí•© ë¶„ì„ ìš”ì•½ (í¸ì§‘ ê°€ëŠ¥)", 
            value=st.session_state.get("summary_content", ""), 
            placeholder="Q&A ë‚´ìš©ì„ ë¶ˆëŸ¬ì˜¤ë©´ ê·œì œ/ë¦¬ì½œì‚¬ë¡€ë¥¼ í†µí•©í•˜ì—¬ ë¶„ì„ ìš”ì•½ë©ë‹ˆë‹¤.",
            height=400,
            key="summary_editor",
            help="AIê°€ ê·œì œ/ë¦¬ì½œì‚¬ë¡€ ëª¨ë“  ëŒ€í™”ë¥¼ í†µí•© ë¶„ì„í•œ ìš”ì•½ì…ë‹ˆë‹¤. í•„ìš”ì‹œ ì§ì ‘ í¸ì§‘ ê°€ëŠ¥í•©ë‹ˆë‹¤."
        )
        
        # í¸ì§‘ëœ ë‚´ìš© ìë™ ì €ì¥
        if edited_summary != st.session_state.get("summary_content", ""):
            st.session_state.summary_content = edited_summary

# show_basic_info_form í•¨ìˆ˜ ëì— ì¶”ê°€
def show_basic_info_form():
    """ê¸°ë³¸ ì •ë³´ ì…ë ¥ í¼ - ìµœì í™”"""
    narrow_col, _ = st.columns([0.8, 0.2])

    with narrow_col:
        render_project_selector()
        st.markdown("---")
        render_product_info_section()
        render_background_section()
        render_risk_summary_section()
        render_summary_display()  # ìš”ì•½ í‘œì‹œ ì¶”ê°€

def get_korean_datetime():
    """í•œêµ­ ì‹œê°„ëŒ€ ê¸°ì¤€ í˜„ì¬ ë‚ ì§œ/ì‹œê°„ ë°˜í™˜"""
    try:
        # í•œêµ­ ì‹œê°„ëŒ€ ì„¤ì •
        korean_tz = pytz.timezone('Asia/Seoul')
        
        # UTC í˜„ì¬ ì‹œê°„ì„ í•œêµ­ ì‹œê°„ìœ¼ë¡œ ë³€í™˜
        utc_now = datetime.utcnow()
        utc_now = pytz.utc.localize(utc_now)
        korean_now = utc_now.astimezone(korean_tz)
        
        return korean_now
    except:
        # pytzê°€ ì—†ê±°ë‚˜ ì˜¤ë¥˜ ì‹œ UTC+9 ìˆ˜ë™ ê³„ì‚°
        utc_now = datetime.utcnow()
        korean_now = utc_now + timedelta(hours=9)
        return korean_now

def create_excel_report():
    """openpyxlì„ ì‚¬ìš©í•œ ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„±"""
    try:
        template_path = './components/genai_rpa.xlsx'
        
        # í•œêµ­ ì‹œê°„ ê¸°ì¤€ìœ¼ë¡œ ë‚ ì§œ ìƒì„±
        korean_now = get_korean_datetime()
        timestamp = korean_now.strftime('%Y%m%d_%H%M%S')
        current_date = korean_now.strftime('%Yë…„ %mì›” %dì¼')
        
        # í…œí”Œë¦¿ íŒŒì¼ì´ ì—†ëŠ” ê²½ìš° ìƒˆë¡œ ìƒì„±
        if not os.path.exists(template_path):
            return create_excel_report_from_scratch(timestamp, current_date)

        output_filename = f"ë¶„ì„ë¦¬í¬íŠ¸_{timestamp}.xlsx"

        # íŒŒì¼ ë³µì‚¬
        shutil.copy(template_path, output_filename)

        # openpyxlë¡œ ì—‘ì…€ ì²˜ë¦¬
        wb = load_workbook(output_filename)
        ws = wb.active  # ì²« ë²ˆì§¸ ì›Œí¬ì‹œíŠ¸ ì„ íƒ
        
        # ë°ì´í„° ì…ë ¥ (ì…€ ì£¼ì†ŒëŠ” í…œí”Œë¦¿ì— ë§ê²Œ ì¡°ì •)
        ws['E8'] = st.session_state.get("product_name", "")
        ws['E10'] = st.session_state.get("target_name", "")
        ws['E12'] = st.session_state.get("background", "")
        ws['E19'] = st.session_state.get("summary_content", "")
        ws['J6'] = current_date  # í•œêµ­ ì‹œê°„ ê¸°ì¤€ ë‚ ì§œ
        ws['C4'] = f"{st.session_state.get('product_name', '')} ìš”ì•½ ë¦¬í¬íŠ¸"

        # íŒŒì¼ ì €ì¥
        wb.save(output_filename)
        wb.close()
        
        return True, output_filename

    except Exception as e:
        return False, f"ì—‘ì…€ íŒŒì¼ ìƒì„± ì¤‘ ì˜¤ë¥˜: {str(e)}"

def create_excel_report_from_scratch(timestamp=None, current_date=None):
    """í…œí”Œë¦¿ì´ ì—†ì„ ë•Œ ì²˜ìŒë¶€í„° ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„± - ì‹œê°„ëŒ€ ìˆ˜ì •"""
    try:
        from openpyxl import Workbook
        
        # í•œêµ­ ì‹œê°„ ê¸°ì¤€ìœ¼ë¡œ ë‚ ì§œ ìƒì„± (íŒŒë¼ë¯¸í„°ê°€ ì—†ì„ ê²½ìš°)
        if not timestamp or not current_date:
            korean_now = get_korean_datetime()
            timestamp = korean_now.strftime('%Y%m%d_%H%M%S')
            current_date = korean_now.strftime('%Yë…„ %mì›” %dì¼')
        
        output_filename = f"ë¶„ì„ë¦¬í¬íŠ¸_{timestamp}.xlsx"
        
        # ìƒˆ ì›Œí¬ë¶ ìƒì„±
        wb = Workbook()
        ws = wb.active
        ws.title = "ìˆ˜ì¶œ ì œì•ˆì„œ ë¶„ì„ ë¦¬í¬íŠ¸"
        
        # ìŠ¤íƒ€ì¼ ì •ì˜
        header_font = Font(name='ë§‘ì€ ê³ ë”•', size=14, bold=True)
        title_font = Font(name='ë§‘ì€ ê³ ë”•', size=16, bold=True)
        normal_font = Font(name='ë§‘ì€ ê³ ë”•', size=10)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # ì œëª© ë° í—¤ë” ì„¤ì •
        ws['A1'] = "ìˆ˜ì¶œ ì œì•ˆì„œ ë¶„ì„ ë¦¬í¬íŠ¸"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f"ìƒì„±ì¼: {current_date}"  # í•œêµ­ ì‹œê°„ ê¸°ì¤€ ë‚ ì§œ
        ws['A2'].font = normal_font
        
        # ì œí’ˆ ì •ë³´ ì„¹ì…˜
        row = 4
        ws[f'A{row}'] = "ğŸ“¦ ì œí’ˆ ì •ë³´"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:J{row}')
        
        row += 1
        ws[f'A{row}'] = "ì œí’ˆëª…:"
        ws[f'B{row}'] = st.session_state.get("product_name", "")
        
        row += 1
        ws[f'A{row}'] = "íƒ€ê²Ÿì¸µ:"
        ws[f'B{row}'] = st.session_state.get("target_name", "")
        
        # ì¶”ì§„ ë°°ê²½ ì„¹ì…˜
        row += 2
        ws[f'A{row}'] = "ğŸ¯ ì¶”ì§„ ë°°ê²½"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:J{row}')
        
        row += 1
        background_text = st.session_state.get("background", "")
        ws[f'A{row}'] = background_text
        ws.merge_cells(f'A{row}:J{row+5}')  # ë°°ê²½ ì„¤ëª…ì„ ìœ„í•œ í° ì…€
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # ê·œì œ ë¦¬ìŠ¤í¬ ìš”ì•½ ì„¹ì…˜
        row += 7
        ws[f'A{row}'] = "âš ï¸ ê·œì œ ë¦¬ìŠ¤í¬ ìš”ì•½"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:J{row}')
        
        row += 1
        summary_text = st.session_state.get("summary_content", "")
        ws[f'A{row}'] = summary_text
        ws.merge_cells(f'A{row}:J{row+10}')  # ìš”ì•½ì„ ìœ„í•œ í° ì…€
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # ì—´ ë„ˆë¹„ ì¡°ì •
        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 12
        
        # íŒŒì¼ ì €ì¥
        wb.save(output_filename)
        wb.close()
        
        return True, output_filename
        
    except Exception as e:
        return False, f"ì—‘ì…€ íŒŒì¼ ìƒì„± ì¤‘ ì˜¤ë¥˜: {str(e)}"

def add_excel_export_button():
    """ì—‘ì…€ ë‚´ë³´ë‚´ê¸° ë²„íŠ¼ - openpyxl ë²„ì „"""
    
    # í•„ìˆ˜ ë°ì´í„° ì²´í¬
    required_fields = ["product_name", "target_name", "background"]
    has_required_data = all(st.session_state.get(field, "") for field in required_fields)
    
    # ì²˜ë¦¬ ìƒíƒœ ì²´í¬
    is_processing = st.session_state.get("ai_processing", False)
    button_disabled = not has_required_data or is_processing
    
    if not has_required_data:
        st.warning("âš ï¸ ì œí’ˆëª…, íƒ€ê²Ÿì¸µ, ì¶”ì§„ë°°ê²½ì„ ëª¨ë‘ ì…ë ¥í•´ì£¼ì„¸ìš”.")
    
    if is_processing:
        st.info("ğŸ”„ AI ë¶„ì„ ì²˜ë¦¬ ì¤‘ì…ë‹ˆë‹¤...")
    
    # ì—‘ì…€ ìƒì„± ë²„íŠ¼
    if st.button(
        "ğŸ“Š í†µí•© ë¶„ì„ ë¦¬í¬íŠ¸ ìƒì„± (Excel)", 
        use_container_width=True,
        disabled=button_disabled,
        help="ì…ë ¥ëœ ì •ë³´ì™€ í†µí•© ë¶„ì„ ê²°ê³¼ë¥¼ ë°”íƒ•ìœ¼ë¡œ ì—‘ì…€ ë¦¬í¬íŠ¸ë¥¼ ìƒì„±í•©ë‹ˆë‹¤."
    ):
        with st.spinner("ğŸ“ ì—‘ì…€ ë¦¬í¬íŠ¸ ìƒì„± ì¤‘..."):
            success, result = create_excel_report()
            
            if success:
                st.success(f"âœ… ë¦¬í¬íŠ¸ ìƒì„± ì™„ë£Œ!")
                st.info(f"ğŸ“ íŒŒì¼ëª…: {result}")
                
                # ë‹¤ìš´ë¡œë“œ ë²„íŠ¼
                try:
                    with open(result, "rb") as file:
                        st.download_button(
                            label="ğŸ“¥ ì—‘ì…€ íŒŒì¼ ë‹¤ìš´ë¡œë“œ",
                            data=file.read(),
                            file_name=result,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    # ì„ì‹œ íŒŒì¼ ì •ë¦¬
                    try:
                        os.remove(result)
                    except:
                        pass
                        
                except Exception as e:
                    st.error(f"ë‹¤ìš´ë¡œë“œ ì¤€ë¹„ ì¤‘ ì˜¤ë¥˜: {e}")
            else:
                st.error(f"âŒ {result}")

# ê¸°íƒ€ í•¨ìˆ˜ë“¤ì€ ê¸°ì¡´ ì½”ë“œ ìœ ì§€í•˜ë˜ í•„ìš”ì‹œ ìµœì í™”
def show_product_analysis():
    """ì œí’ˆ ë¶„ì„ ì„¹ì…˜ - ê¸°ì¡´ ì½”ë“œ ìœ ì§€"""
    pass

def show_report_generation():
    """ì œì•ˆì„œ ìƒì„± ì„¹ì…˜ - ê¸°ì¡´ ì½”ë“œ ìœ ì§€"""
    pass

def show_results_section():
    """ê²°ê³¼ í‘œì‹œ ì„¹ì…˜ - ê¸°ì¡´ ì½”ë“œ ìœ ì§€"""
    pass
